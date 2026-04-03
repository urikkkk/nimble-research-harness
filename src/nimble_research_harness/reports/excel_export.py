"""McKinsey-level Excel export — two tiers based on budget/data size.

Under 10m budget: Fixed 5-tab template (no LLM, instant).
10m+ budget: LLM analyzes evidence structure and generates custom layout.
"""

from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# --- Style constants ---
NAVY = "1B3A5C"
DBLUE = "2B5797"
MBLUE = "4472C4"
LBLUE = "D6E4F0"
XLBLUE = "EDF2F9"
WHITE = "FFFFFF"
GRAY = "808080"
GREEN = "548235"
RED = "C00000"
ORANGE = "ED7D31"

HDR_FONT = Font(bold=True, color=WHITE, size=11, name="Calibri")
HDR_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
ALT_FILL = PatternFill(start_color=XLBLUE, end_color=XLBLUE, fill_type="solid")
SEC_FILL = PatternFill(start_color=LBLUE, end_color=LBLUE, fill_type="solid")
SEC_FONT = Font(bold=True, size=11, color=NAVY, name="Calibri")
BODY_FONT = Font(size=10, name="Calibri")
BOLD_FONT = Font(size=10, name="Calibri", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(bottom=Side(style="thin", color="D0D0D0"))


def _header_row(ws, cols: list[str], row: int = 1) -> None:
    for i, col in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=col)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = CENTER
    ws.freeze_panes = f"A{row + 1}"
    ws.auto_filter.ref = ws.dimensions


def _set_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def export_excel(
    output_path: str,
    user_query: str,
    report: dict[str, Any],
    claims: list[dict[str, Any]],
    evidence: list[dict[str, Any]],
    session_meta: Optional[dict[str, Any]] = None,
) -> str:
    """Generate McKinsey-level Excel from research session data.

    Uses a fixed 5-tab template for fast, consistent output.
    """
    wb = Workbook()

    domains = Counter(e.get("source_domain", "") for e in evidence)
    meta = session_meta or {}

    # === TAB 1: EXECUTIVE DASHBOARD ===
    ws1 = wb.active
    ws1.title = "Executive Dashboard"
    ws1.sheet_properties.tabColor = NAVY

    ws1.merge_cells("A1:H1")
    c = ws1["A1"]
    c.value = (report.get("title") or "Research Intelligence Report").upper()
    c.font = Font(bold=True, size=16, color=WHITE, name="Calibri")
    c.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    c.alignment = CENTER

    ws1.merge_cells("A2:H2")
    c = ws1["A2"]
    c.value = user_query[:200]
    c.font = Font(size=11, color=WHITE, italic=True, name="Calibri")
    c.fill = PatternFill(start_color=DBLUE, end_color=DBLUE, fill_type="solid")
    c.alignment = CENTER

    ws1.merge_cells("A3:H3")
    c = ws1["A3"]
    budget = meta.get("time_budget", "")
    elapsed = meta.get("elapsed_seconds", 0)
    c.value = f"Evidence: {len(evidence)} items | Domains: {len(domains)} | Claims: {len(claims)} | Budget: {budget} | Elapsed: {elapsed:.0f}s"
    c.font = Font(size=9, color=GRAY, name="Calibri")
    c.alignment = CENTER

    # Executive summary
    r = 5
    ws1.merge_cells(f"A{r}:H{r}")
    ws1.cell(row=r, column=1, value="EXECUTIVE SUMMARY").font = SEC_FONT
    ws1.cell(row=r, column=1).fill = SEC_FILL
    r += 1
    ws1.merge_cells(f"A{r}:H{r}")
    ws1.cell(row=r, column=1, value=report.get("executive_summary", "")).font = BODY_FONT
    ws1.cell(row=r, column=1).alignment = WRAP

    # Key findings
    r += 2
    ws1.merge_cells(f"A{r}:H{r}")
    ws1.cell(row=r, column=1, value="KEY FINDINGS").font = SEC_FONT
    ws1.cell(row=r, column=1).fill = SEC_FILL
    r += 1
    for finding in report.get("key_findings", []):
        ws1.merge_cells(f"A{r}:H{r}")
        ws1.cell(row=r, column=1, value=f"• {finding}").font = BODY_FONT
        ws1.cell(row=r, column=1).alignment = WRAP
        r += 1

    # Top sources
    r += 1
    ws1.merge_cells(f"A{r}:H{r}")
    ws1.cell(row=r, column=1, value="TOP SOURCES").font = SEC_FONT
    ws1.cell(row=r, column=1).fill = SEC_FILL
    r += 1
    for domain, count in domains.most_common(10):
        ws1.cell(row=r, column=1, value=domain).font = BOLD_FONT
        ws1.cell(row=r, column=2, value=count).font = Font(bold=True, size=12, color=NAVY, name="Calibri")
        r += 1

    _set_widths(ws1, [20, 12, 16, 16, 16, 16, 16, 16])

    # === TAB 2: VERIFIED CLAIMS ===
    ws2 = wb.create_sheet("Verified Claims")
    ws2.sheet_properties.tabColor = GREEN
    cols2 = ["#", "Claim", "Confidence", "Category", "Importance", "Source URLs", "Key Excerpts"]
    _header_row(ws2, cols2)

    for i, c in enumerate(claims, 1):
        r = i + 1
        ws2.cell(row=r, column=1, value=i).font = BODY_FONT
        ws2.cell(row=r, column=2, value=c.get("statement", "")).font = BODY_FONT
        ws2.cell(row=r, column=2).alignment = WRAP
        conf = c.get("confidence", "")
        ws2.cell(row=r, column=3, value=conf).font = Font(
            color=GREEN if "verified" in str(conf).lower() else GRAY,
            bold=True, size=10, name="Calibri",
        )
        ws2.cell(row=r, column=4, value=c.get("category", "")).font = BODY_FONT
        ws2.cell(row=r, column=5, value=c.get("importance")).font = BODY_FONT
        urls = c.get("source_urls", [])
        ws2.cell(row=r, column=6, value="\n".join(urls[:5])).font = Font(size=8, color=MBLUE, name="Calibri")
        ws2.cell(row=r, column=6).alignment = WRAP
        excerpts = c.get("excerpts", [])
        ws2.cell(row=r, column=7, value="\n---\n".join(e[:200] for e in excerpts[:3])).font = Font(size=8, name="Calibri")
        ws2.cell(row=r, column=7).alignment = WRAP
        if r % 2 == 0:
            for col in range(1, 8):
                ws2.cell(row=r, column=col).fill = ALT_FILL

    _set_widths(ws2, [5, 60, 18, 16, 10, 40, 50])

    # === TAB 3: EVIDENCE BY SOURCE ===
    ws3 = wb.create_sheet("Evidence by Source")
    ws3.sheet_properties.tabColor = DBLUE

    # Group evidence by domain
    by_domain = defaultdict(list)
    for e in evidence:
        by_domain[e.get("source_domain", "unknown")].append(e)

    cols3 = ["Domain", "Count", "Sample Titles"]
    _header_row(ws3, cols3)
    r = 2
    for domain in sorted(by_domain.keys(), key=lambda d: -len(by_domain[d])):
        items = by_domain[domain]
        titles = "; ".join((e.get("title") or "")[:60] for e in items[:5])
        ws3.cell(row=r, column=1, value=domain).font = BOLD_FONT
        ws3.cell(row=r, column=2, value=len(items)).font = Font(bold=True, size=12, color=NAVY, name="Calibri")
        ws3.cell(row=r, column=2).alignment = CENTER
        ws3.cell(row=r, column=3, value=titles).font = Font(size=8, color=GRAY, name="Calibri")
        ws3.cell(row=r, column=3).alignment = WRAP
        if r % 2 == 0:
            for col in range(1, 4):
                ws3.cell(row=r, column=col).fill = ALT_FILL
        r += 1

    _set_widths(ws3, [30, 10, 80])

    # === TAB 4: RAW EVIDENCE ===
    ws4 = wb.create_sheet("Raw Evidence")
    ws4.sheet_properties.tabColor = GRAY
    cols4 = ["#", "Domain", "Title", "Content", "URL"]
    _header_row(ws4, cols4)

    for i, e in enumerate(evidence[:2000], 1):
        r = i + 1
        ws4.cell(row=r, column=1, value=i).font = BODY_FONT
        ws4.cell(row=r, column=2, value=e.get("source_domain", "")).font = BODY_FONT
        ws4.cell(row=r, column=3, value=(e.get("title") or "")[:150]).font = BODY_FONT
        ws4.cell(row=r, column=4, value=(e.get("content") or "")[:400]).font = Font(size=8, name="Calibri")
        ws4.cell(row=r, column=4).alignment = WRAP
        ws4.cell(row=r, column=5, value=e.get("source_url", "")).font = Font(size=8, color=MBLUE, name="Calibri")
        if i % 2 == 0:
            for col in range(1, 6):
                ws4.cell(row=r, column=col).fill = ALT_FILL

    _set_widths(ws4, [5, 22, 40, 60, 50])

    # === TAB 5: METHODOLOGY ===
    ws5 = wb.create_sheet("Methodology")
    ws5.sheet_properties.tabColor = "808080"

    ws5.merge_cells("A1:D1")
    ws5.cell(row=1, column=1, value="RESEARCH METHODOLOGY").font = SEC_FONT
    ws5.cell(row=1, column=1).fill = SEC_FILL

    method_data = [
        ("Query", user_query),
        ("Budget", meta.get("time_budget", "")),
        ("Elapsed", f"{meta.get('elapsed_seconds', 0):.0f}s"),
        ("Evidence Items", str(len(evidence))),
        ("Unique Domains", str(len(domains))),
        ("Verified Claims", str(len(claims))),
        ("Methodology", report.get("methodology", "")),
    ]
    for i, (label, value) in enumerate(method_data, 2):
        ws5.cell(row=i, column=1, value=label).font = BOLD_FONT
        ws5.cell(row=i, column=2, value=value).font = BODY_FONT
        ws5.cell(row=i, column=2).alignment = WRAP

    if report.get("known_unknowns"):
        r = len(method_data) + 3
        ws5.cell(row=r, column=1, value="KNOWN UNKNOWNS").font = SEC_FONT
        ws5.cell(row=r, column=1).fill = SEC_FILL
        for item in report["known_unknowns"]:
            r += 1
            ws5.cell(row=r, column=1, value=f"• {item}").font = BODY_FONT
            ws5.cell(row=r, column=1).alignment = WRAP

    if report.get("limitations"):
        r += 2
        ws5.cell(row=r, column=1, value="LIMITATIONS").font = SEC_FONT
        ws5.cell(row=r, column=1).fill = SEC_FILL
        for item in report["limitations"]:
            r += 1
            ws5.cell(row=r, column=1, value=f"• {item}").font = BODY_FONT
            ws5.cell(row=r, column=1).alignment = WRAP

    _set_widths(ws5, [20, 80, 20, 20])

    wb.save(output_path)
    return output_path
